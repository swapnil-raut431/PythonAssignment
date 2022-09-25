import openpyxl
from openpyxl.styles import PatternFill

class xlutils:
    def getrowcount(file,sheetname):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        return (sheet.max_row)

    def getcolumncount(file,sheetname):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        return (sheet.max_column)

    def readdata(file,sheetname,rownum,columnno):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        return sheet.cell(row=rownum,column=columnno).value

    def columndata(file,sheetname,rownum,columnno,data):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        sheet.cell(row=rownum,column=columnno).value=data
        workbook.save(file)

    def fillgreencolor(file,sheetname,rownum,columnno):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        greenFill=PatternFill(start_color='60b212',
                              end_color='60b212',
                              fill_type='solid')
        sheet.cell(rownum,columnno).fill=greenFill
        workbook.save(file)

    def fillredcolor(file,sheetname,rownum,columnno):
        workbook=openpyxl.load_workbook(file)
        sheet=workbook[sheetname]
        redFill=PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
        sheet.cell(rownum,columnno).fill=redFill
        workbook.save(file)

